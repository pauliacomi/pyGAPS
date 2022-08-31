"""Parse SMS DVS xlsx output files."""
import dateutil.parser
import openpyxl

from . import utils as util
from .unit_parsing import parse_temperature_string

_META_DICT = {
    'material': {
        'text': ('sample name:', ),
        'type': 'string',
        "xl_ref": (0, 1)  # row, column
    },
    'material_mass': {
        'text': ('ref. mass', ),
        'type': 'numeric',
        "xl_ref": (0, 1)
    },
    'material_mass_initial': {
        'text': ('initial mass [mg]:', ),  # TODO not only mg?
        'type': 'numeric',
        "xl_ref": (0, 1)
    },
    'adsorbate': {
        'text': ('vapour:', ),
        'type': 'string',
        "xl_ref": (0, 1)
    },
    'adsorbate_saturation_pressure': {
        'text': ('vapour pressure [torr]:', ),
        'type': 'numeric',
        "xl_ref": (0, 1)
    },
    'operator': {
        'text': ('user name:', ),
        'type': 'string',
        "xl_ref": (0, 1)
    },
    'date': {
        'text': ('raw data file created:', ),
        'type': 'date',
        "xl_ref": (0, 1)
    },
    'dvs_file_version': {
        'text': ('file version:', ),
        'type': 'string',
        "xl_ref": (0, 1)
    },
    'dvs_sequence_name': {
        'text': ('sequence name:', ),
        'type': 'string',
        "xl_ref": (0, 1)
    },
    'dvs_sequence_creation_date': {
        'text': ('sequence created:', ),
        'type': 'date',
        "xl_ref": (0, 1)
    },
    'dvs_method_name': {
        'text': ('method name:', ),
        'type': 'string',
        "xl_ref": (0, 1)
    },
    'dvs_method_creation_date': {
        'text': ('method created:', ),
        'type': 'date',
        "xl_ref": (0, 1)
    },
    'dvs_method_modified': {
        'text': ('method modified:', ),
        'type': 'numeric',
        "xl_ref": (0, 1)
    },
    'dvs_sample_number': {
        'text': ('sample number:', ),
        'type': 'numeric',
        "xl_ref": (0, 1)
    },
    'dvs_sample_description': {
        'text': ('sample description:', ),
        'type': 'string',
        "xl_ref": (0, 1)
    },
    'dvs_data_save_interval': {
        'text': ('data saving interval [seconds]:', ),
        'type': 'numeric',
        "xl_ref": (0, 1)
    },
    'dvs_control_mode': {
        'text': ('control mode', ),
        'type': 'string',
        "xl_ref": (0, 1)
    },
    'dvs_material_mass_determination': {
        'text': ('ref. mass option', ),
        'type': 'string',
        "xl_ref": (0, 1)
    },
    'dvs_imported_by_dll': {
        'text': ('importbydllseries', ),
        'type': 'string',
        "xl_ref": (0, 1)
    },
}

_DATA_DICT = {}


def parse(path):
    """
    Parse an xlsx file analysed through SMS DVS software
    to obtain the isotherm.

    Parameters
    ----------
    path: str
        The location of a processed isotherm in Excel.

    Returns
    -------
    dict
        A dictionary containing isotherm information.
    """
    meta = {}
    data = {}

    # open the workbook
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # local for efficiency
    meta_dict = _META_DICT.copy()

    # First get metadata/kinetics
    rawdata_sheet = workbook['DVS Data']
    # we know data is left-aligned
    # so we only iterate rows
    for row in rawdata_sheet.rows:

        # if first cell is not filled -> blank row
        cell_value = row[0]
        if not cell_value.value:
            continue

        # We do not take kinetics for the moment
        if cell_value.value == "Time [minutes]":  # If "kinetic data" section
            break

        key = cell_value.value.lower()
        try:
            key = util.search_key_in_def_dict(key, meta_dict)
        except StopIteration:
            continue

        ref = meta_dict[key]['xl_ref']
        tp = meta_dict[key]['type']
        del meta_dict[key]  # delete for efficiency

        # handle different data types
        val = rawdata_sheet.cell(cell_value.row + ref[0], cell_value.column + ref[1]).value
        if val == '':
            meta[key] = None
        elif tp == 'numeric':
            meta[key] = val
        elif tp == 'string':
            meta[key] = util.handle_excel_string(val)
        elif tp == 'date':
            meta[key] = _handle_dvs_date(val)

    # Then get data and some remaining metadata
    book = None
    if "Iso Report" in workbook:
        book = "Iso Report"
    elif "Iso Report (Torr)" in workbook:
        book = "Iso Report (Torr)"
    else:
        raise Exception("Could not find a processed isotherm in the file.")

    iso_sheet = workbook[book]

    # data is randomly distributed
    # all has to be iterated
    for row in iso_sheet.rows:
        for cell in row:
            if not cell.value:
                continue

            if cell.value == "Temp:":
                temp = iso_sheet.cell(cell.row, cell.column + 1).value
                comp = temp.split()
                meta["temperature"] = float(comp[0])
                meta["temperature_unit"] = parse_temperature_string(comp[1])

            elif cell.value == "Cycle 1":
                head, unit = _parse_header(iso_sheet, cell.row - 2, cell.column + 1)
                meta.update(unit)
                data = _parse_data(iso_sheet, cell.row, head)
                data = _sort_data(data, head)

                # TODO other cycles
                # Finished for now
                break

    # Set extra metadata
    meta["material_mass_unit"] = "mg"
    meta["material_basis"] = "mass"
    meta["material_unit"] = None

    return meta, data


def _parse_header(sheet, row, col):
    """
    Parse a header DVS header.

    Starts at the upper left corner (at "Target").

    """

    headers = {}
    units = {}

    # determine pressure mode
    pressure_mode = sheet.cell(row + 1, col).value
    if pressure_mode == "% P/Po":
        units["pressure_mode"] = "relative%"
        units["pressure_unit"] = None
    elif pressure_mode == "Pres.":
        units["pressure_mode"] = "absolute"
        units["pressure_unit"] = "torr"
    headers["pressure_target"] = col

    # determine target/actual display
    pressure_output = sheet.cell(row, col + 1).value
    if pressure_output == "Actual":
        headers["pressure_actual_ads"] = col + 1
        headers["loading_ads"] = col + 2
        headers["pressure_actual_des"] = col + 3
        headers["loading_des"] = col + 4
    else:
        headers["loading_ads"] = col + 1
        headers["loading_des"] = col + 2

    # TODO does this change?
    units["loading_basis"] = "percent"
    units["loading_unit"] = None

    return headers, units


def _parse_data(sheet, row, head):
    """Return start and stop points for adsorption and desorption."""

    data = {k: [] for k in head}

    # pressure_target column
    col = head["pressure_target"]
    while True:
        text = sheet.cell(row, col).value
        if text is None:
            break

        rowdata = sheet[row]
        for key, kcol in head.items():
            data[key].append(rowdata[kcol - 1].value)

        row += 1

    return data


def _sort_data(data, head):

    ds = {}
    ds["branch"] = ([0] * len(data["pressure_target"]) + [1] * len(data["pressure_target"]))
    ds["loading"] = (data["loading_ads"] + data["loading_des"][::-1])

    if "pressure_actual_ads" in head:
        ds["pressure_target"] = (data["pressure_target"] + data["pressure_target"][::-1])
        ds["pressure"] = (data["pressure_actual_ads"] + data["pressure_actual_des"][::-1])
        indices = [i for i, v in enumerate(ds["pressure"]) if v is None]
    else:
        ds["pressure"] = (data["pressure_target"] + data["pressure_target"][::-1])
        indices = [i for i, v in enumerate(ds["loading"]) if v is None]

    # remove empty data
    indices = indices[::-1]
    for v in ds.values():
        for index in indices:
            del v[index]

    return ds


def _handle_dvs_date(text):
    if text == "N/A":
        return None
    return dateutil.parser.parse(text.replace(" UTC ", "")).isoformat()
